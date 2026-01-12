"""
Routes/Blueprints untuk PonPay
"""
from flask import Blueprint, render_template, request, redirect, url_for, g, session, send_file, current_app, flash, jsonify
from db import (query_db, execute_db, get_dashboard_stats, get_monthly_stats, get_category_stats,
                get_all_students, get_student, get_student_payments, get_student_payment_stats,
                add_student, update_student, delete_student, record_history, get_history,
                get_all_users, get_user, create_user, update_user, delete_user, set_user_password, get_user_by_username,
                get_all_bills, create_bill, get_bill, update_bill, delete_bill, mark_bill_paid, get_student_unpaid_amount, get_bill_stats_by_class)
from werkzeug.security import check_password_hash
from datetime import datetime, timedelta
import json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
import os
from fpdf import FPDF
from werkzeug.utils import secure_filename
from functools import wraps
from utils.validation import (
    validate_username, validate_password, validate_email, validate_amount,
    validate_date, validate_name, validate_nisn, validate_phone, validate_kelas,
    validate_jenis_kelamin, validate_alamat, validate_category, validate_description,
    validate_file_upload, validate_student_data, validate_transaction_data,
    validate_user_data, ValidationError, check_rate_limit, flash_validation_errors
)

def _is_admin():
    return session.get('role') == 'admin'


def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not _is_admin():
            flash('Anda tidak memiliki akses ke halaman ini', 'danger')
            return redirect(url_for('dashboard.index'))
        return f(*args, **kwargs)
    return decorated_function

# ============ BLUEPRINTS ============

# Home route - redirect ke login atau dashboard
def create_home_routes(app):
    """Create home routes for the app"""
    @app.route('/')
    def index():
        """Home page - always redirect to login (root should lead to login)"""
        return redirect(url_for('auth.login'))

# Auth Blueprint
auth_bp = Blueprint('auth', __name__)

@auth_bp.route('/login', methods=['GET', 'POST'])
def login():
    """Halaman login"""
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = get_user_by_username(username)
        if user:
            stored_pw = user['password']
            # Normal path: stored password is a hash
            if check_password_hash(stored_pw, password):
                session['user_id'] = user['id']
                session['username'] = user['username']
                # sqlite3.Row does not implement .get(), so access safely
                try:
                    full_name = user['full_name'] if user['full_name'] else user['username']
                except Exception:
                    full_name = user['username']
                try:
                    role = user['role'] if user['role'] else 'user'
                except Exception:
                    role = 'user'
                try:
                    profile_picture = user['profile_picture'] if user['profile_picture'] else None
                except Exception:
                    profile_picture = None
                session['full_name'] = full_name
                session['role'] = role
                session['profile_picture'] = profile_picture
                session['profile_picture_timestamp'] = int(datetime.now().timestamp())
                return redirect(url_for('dashboard.index'))
            # Backwards compatibility: stored password might be plaintext in older DBs
            if stored_pw == password:
                try:
                    # migrate to hashed password
                    set_user_password(user['id'], password)
                except Exception:
                    pass
                session['user_id'] = user['id']
                session['username'] = user['username']
                try:
                    full_name = user['full_name'] if user['full_name'] else user['username']
                except Exception:
                    full_name = user['username']
                try:
                    role = user['role'] if user['role'] else 'user'
                except Exception:
                    role = 'user'
                session['full_name'] = full_name
                session['role'] = role
                return redirect(url_for('dashboard.index'))

        return render_template('login.html', error='Username atau password salah')
    
    return render_template('login.html')

@auth_bp.route('/logout')
def logout():
    """Logout"""
    session.clear()
    return redirect(url_for('auth.login'))

# Dashboard Blueprint
dashboard_bp = Blueprint('dashboard', __name__, url_prefix='/dashboard')

@dashboard_bp.route('/')
def index():
    """Halaman dashboard utama"""
    user_id = session.get('user_id', 1)
    
    stats = get_dashboard_stats(user_id)
    monthly_data = get_monthly_stats(user_id, months=1)
    
    # Format data untuk Chart.js
    months_list = []
    income_data = []
    expense_data = []
    
    today = datetime.now()
    for i in range(11, -1, -1):
        date = today - timedelta(days=30*i)
        month_str = date.strftime('%Y-%m')
        months_list.append(date.strftime('%b'))
        
        # Cari data untuk bulan ini
        income = 0
        expense = 0
        for row in monthly_data:
            if row['month'] == month_str:
                if row['type'] == 'income':
                    income = row['total']
                else:
                    expense = row['total']
        
        income_data.append(income)
        expense_data.append(expense)
    
    return render_template('dashboard.html',
                         stats=stats,
                         months=json.dumps(months_list),
                         income_data=json.dumps(income_data),
                         expense_data=json.dumps(expense_data))

# Transaction Blueprint
transaction_bp = Blueprint('transaction', __name__, url_prefix='/transaction')

@transaction_bp.route('/')
def index():
    """Daftar transaksi"""
    user_id = session.get('user_id', 1)
    
    # Filter
    filter_type = request.args.get('type', 'all')
    filter_category = request.args.get('category', 'all')
    filter_month = request.args.get('month', '')
    
    # Query dengan JOIN ke students untuk menampilkan nama santri
    query = '''
        SELECT t.*,
               COALESCE(s.name, '') as student_name,
               COALESCE(s.nisn, '') as student_nisn
        FROM transactions t
        LEFT JOIN students s ON t.student_id = s.id
        WHERE t.user_id = ?
    '''
    params = [user_id]

    if filter_type != 'all':
        query += 'AND t.type = ? '
        params.append(filter_type)

    if filter_category != 'all':
        query += 'AND t.category = ? '
        params.append(filter_category)

    if filter_month:
        query += 'AND t.date LIKE ? '
        params.append(f'{filter_month}%')
    
    query += 'ORDER BY t.date DESC, t.created_at DESC'
    
    transactions = query_db(query, params)
    
    # Ambil kategori unik untuk dropdown
    categories_income = query_db(
        'SELECT DISTINCT category FROM transactions WHERE user_id = ? AND type = "income" ORDER BY category',
        (user_id,)
    )
    categories_expense = query_db(
        'SELECT DISTINCT category FROM transactions WHERE user_id = ? AND type = "expense" ORDER BY category',
        (user_id,)
    )
    
    return render_template('transaction.html',
                         transactions=transactions,
                         categories_income=[r['category'] for r in categories_income],
                         categories_expense=[r['category'] for r in categories_expense],
                         filter_type=filter_type,
                         filter_category=filter_category,
                         filter_month=filter_month)

@transaction_bp.route('/add', methods=['GET', 'POST'])
def add():
    """Tambah transaksi"""
    user_id = session.get('user_id', 1)
    
    if request.method == 'POST':
        trans_type = request.form.get('type')
        category = request.form.get('category')
        amount = int(request.form.get('amount', 0))
        description = request.form.get('description')
        date = request.form.get('date')
        student_id = request.form.get('student_id') or None
        
        # Convert student_id to int or None
        if student_id:
            try:
                student_id = int(student_id)
            except (ValueError, TypeError):
                student_id = None
        
        # Insert transaksi dengan student_id
        trans_id = execute_db(
            '''INSERT INTO transactions (user_id, student_id, type, category, amount, description, date)
               VALUES (?, ?, ?, ?, ?, ?, ?)''',
            (user_id, student_id, trans_type, category, amount, description, date)
        )

        # Update saldo wallet
        wallet = query_db('SELECT balance FROM wallet WHERE user_id = ?', (user_id,), one=True)
        current_balance = wallet['balance'] if wallet else 0

        if trans_type == 'income':
            new_balance = current_balance + amount
        else:
            new_balance = current_balance - amount

        execute_db('UPDATE wallet SET balance = ? WHERE user_id = ?', (new_balance, user_id))
        try:
            record_history(user_id, 'create', 'transaction', trans_id, f"{category}:{amount}")
        except Exception:
            pass

        return redirect(url_for('transaction.index'))
    
    # Get all students untuk dropdown
    students = get_all_students()
    # Convert sqlite3.Row to plain dicts so templates render cleanly (avoid printing Row repr)
    students = [dict(s) for s in students] if students else []

    return render_template('add_transaction.html', students=students)


@transaction_bp.route('/export-excel')
def export_excel():
    """Export transaksi ke Excel dengan filter yang sama seperti halaman daftar"""
    user_id = session.get('user_id', 1)

    filter_type = request.args.get('type', 'all')
    filter_category = request.args.get('category', 'all')
    filter_month = request.args.get('month', '')

    query = '''
        SELECT t.*, COALESCE(s.name, '') as student_name, COALESCE(s.nisn, '') as student_nisn
        FROM transactions t
        LEFT JOIN students s ON t.student_id = s.id
        WHERE t.user_id = ?
    '''
    params = [user_id]
    if filter_type != 'all':
        query += ' AND t.type = ? '
        params.append(filter_type)
    if filter_category != 'all':
        query += ' AND t.category = ? '
        params.append(filter_category)
    if filter_month:
        query += ' AND t.date LIKE ? '
        params.append(f'{filter_month}%')

    query += ' ORDER BY t.date DESC, t.created_at DESC'

    rows = query_db(query, params)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Transaksi'

    # Styles
    header_fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_alignment = Alignment(horizontal='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ['No.', 'Tanggal', 'Kategori', 'Keterangan', 'Santri', 'NISN', 'Nominal', 'Jenis', 'Created At']
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    for idx, r in enumerate(rows, 1):
        typ = r['type']
        amount = r['amount']
        ws.append([
            idx,
            r['date'],
            r['category'],
            r['description'],
            r['student_name'],
            r['student_nisn'],
            amount,
            typ,
            r.get('created_at') if isinstance(r, dict) else r['created_at']
        ])

    # Column widths
    widths = [5, 15, 25, 40, 25, 15, 15, 12, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'transaksi_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

@transaction_bp.route('/edit/<int:id>', methods=['GET', 'POST'])
def edit(id):
    """Edit transaksi"""
    user_id = session.get('user_id', 1)
    trans = query_db('SELECT * FROM transactions WHERE id = ? AND user_id = ?', (id, user_id), one=True)
    
    if not trans:
        return redirect(url_for('transaction.index'))
    
    if request.method == 'POST':
        trans_type = request.form.get('type')
        category = request.form.get('category')
        amount = int(request.form.get('amount', 0))
        description = request.form.get('description')
        date = request.form.get('date')
        student_id = request.form.get('student_id') or None
        
        # Convert student_id to int or None
        if student_id:
            try:
                student_id = int(student_id)
            except (ValueError, TypeError):
                student_id = None
        
        # Hitung perubahan saldo
        wallet = query_db('SELECT balance FROM wallet WHERE user_id = ?', (user_id,), one=True)
        current_balance = wallet['balance'] if wallet else 0

        # Kembalikan nilai lama terlebih dahulu
        if trans['type'] == 'income':
            current_balance -= trans['amount']
        else:
            current_balance += trans['amount']

        # Terapkan nilai baru
        if trans_type == 'income':
            current_balance += amount
        else:
            current_balance -= amount

        execute_db('UPDATE wallet SET balance = ? WHERE user_id = ?', (current_balance, user_id))

        # Update transaksi dengan student_id
        execute_db(
            '''UPDATE transactions SET student_id = ?, type = ?, category = ?, amount = ?, description = ?, date = ?
               WHERE id = ?''',
            (student_id, trans_type, category, amount, description, date, id)
        )
        try:
            record_history(user_id, 'update', 'transaction', id, f"{category}:{amount}")
        except Exception:
            pass

        return redirect(url_for('transaction.index'))
    
    # Get all students untuk dropdown
    # Convert transaction and students to plain dicts to avoid Jinja printing sqlite Row objects
    try:
        trans = dict(trans)
    except Exception:
        # if already a dict or mapping, keep as-is
        pass

    students = get_all_students()
    students = [dict(s) for s in students] if students else []

    return render_template('edit_transaction.html', transaction=trans, students=students)

@transaction_bp.route('/delete/<int:id>')
def delete(id):
    """Hapus transaksi"""
    user_id = session.get('user_id', 1)
    trans = query_db('SELECT * FROM transactions WHERE id = ? AND user_id = ?', (id, user_id), one=True)
    
    if trans:
        # Update saldo
        wallet = query_db('SELECT balance FROM wallet WHERE user_id = ?', (user_id,), one=True)
        current_balance = wallet['balance'] if wallet else 0

        if trans['type'] == 'income':
            current_balance -= trans['amount']
        else:
            current_balance += trans['amount']

        execute_db('UPDATE wallet SET balance = ? WHERE user_id = ?', (current_balance, user_id))

        # Hapus transaksi
        execute_db('DELETE FROM transactions WHERE id = ?', (id,))
        try:
            record_history(user_id, 'delete', 'transaction', id, None)
        except Exception:
            pass
    
    return redirect(url_for('transaction.index'))

@transaction_bp.route('/receipt/<int:id>')
def receipt(id):
    """Generate kwitansi PDF untuk transaksi tunggal"""
    user_id = session.get('user_id', 1)
    
    # Ambil data transaksi dengan info santri
    query = '''
        SELECT t.*, s.name as student_name, s.nisn as student_nisn, s.kelas as student_kelas
        FROM transactions t
        LEFT JOIN students s ON t.student_id = s.id
        WHERE t.id = ? AND t.user_id = ?
    '''
    trans = query_db(query, (id, user_id), one=True)
    
    if not trans:
        flash('Transaksi tidak ditemukan', 'danger')
        return redirect(url_for('transaction.index'))

    # Ambil setting nama pondok
    settings = query_db('SELECT `key`, value FROM settings')
    settings_dict = {row['key']: row['value'] for row in settings}
    pondok_name = settings_dict.get('pondok_name', 'Pondok Pesantren Al Huda')

    # Buat PDF
    pdf = FPDF()
    pdf.add_page()
    
    # Header
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, pondok_name.upper(), 0, 1, "C")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 5, "Sistem Pembayaran Terpadu (PonPay)", 0, 1, "C")
    pdf.ln(5)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(10)

    # Title
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "BUKTI PEMBAYARAN", 0, 1, "C")
    pdf.ln(5)

    # Info Transaksi
    pdf.set_font("Helvetica", "", 11)
    
    col_width = 45
    
    def add_row(label, value):
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(col_width, 8, label, 0, 0)
        pdf.set_font("Helvetica", "", 11)
        pdf.cell(0, 8, f": {value}", 0, 1)

    add_row("No. Transaksi", f"TRX-{trans['id']}")
    add_row("Tanggal", trans['date'])
    add_row("Penerima", session.get('full_name', 'Bendahara'))
    pdf.ln(5)

    # Info Santri
    if trans['student_name']:
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(0, 8, "Data Santri:", 0, 1)
        pdf.set_font("Helvetica", "", 11)
        add_row("Nama", trans['student_name'])
        add_row("NISN", trans['student_nisn'] or "-")
        add_row("Kelas", trans['student_kelas'] or "-")
        pdf.ln(5)

    # Rincian
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 8, "Rincian Pembayaran:", 0, 1)
    pdf.set_font("Helvetica", "", 11)
    
    # Table Header
    pdf.set_fill_color(240, 240, 240)
    pdf.cell(130, 10, "Keterangan / Kategori", 1, 0, "C", True)
    pdf.cell(60, 10, "Jumlah", 1, 1, "C", True)
    
    # Table Content
    description = f"{trans['category']} - {trans['description']}"
    pdf.cell(130, 20, description, 1, 0, "L")
    
    # Format Rupiah manual for PDF
    amount_str = f"Rp {int(trans['amount']):,.0f}".replace(',', '.')
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(60, 20, amount_str, 1, 1, "R")
    
    pdf.ln(20)

    # Tanda Tangan
    current_date = datetime.now().strftime("%d %m %Y")
    pdf.set_font("Helvetica", "", 11)
    pdf.cell(120, 8, "", 0, 0)
    pdf.cell(0, 8, f"Dicetak pada: {current_date}", 0, 1, "C")
    pdf.ln(20)
    pdf.cell(120, 8, "", 0, 0)
    pdf.cell(0, 8, "( ____________________ )", 0, 1, "C")
    pdf.cell(120, 8, "", 0, 0)
    pdf.cell(0, 8, "Bendahara Pondok", 0, 1, "C")

    # Output to BytesIO
    output = BytesIO()
    pdf_content = pdf.output(dest='S')
    output.write(pdf_content)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/pdf',
        as_attachment=False,
        download_name=f"kwitansi_{trans['id']}.pdf"
    )

# Statistics Blueprint
statistics_bp = Blueprint('statistics', __name__, url_prefix='/statistics')

@statistics_bp.route('/')
def index():
    """Halaman statistik"""
    user_id = session.get('user_id', 1)

    filter_period = request.args.get('period', '1')  # 1, 3, 6, 12 bulan
    period_months = int(filter_period)

    # Ambil data kategori
    expense_categories = get_category_stats(user_id, 'expense', period_months)
    income_categories = get_category_stats(user_id, 'income', period_months)

    # Format data untuk chart - konversi Decimal ke float dan handle None
    expense_labels = [r['category'] for r in expense_categories] if expense_categories else []
    expense_values = [float(r['total']) if r['total'] is not None else 0.0 for r in expense_categories] if expense_categories else []

    income_labels = [r['category'] for r in income_categories] if income_categories else []
    income_values = [float(r['total']) if r['total'] is not None else 0.0 for r in income_categories] if income_categories else []

    # Warna untuk chart
    colors = ['#2E7D32', '#43A047', '#66BB6A', '#81C784', '#A5D6A7', '#C8E6C9']

    # Ambil statistik tunggakan per kelas - konversi Decimal ke float dan handle None
    class_stats = get_bill_stats_by_class()
    class_labels = [r['kelas'] for r in class_stats] if class_stats else []
    class_values = [float(r['total_unpaid']) if r['total_unpaid'] is not None else 0.0 for r in class_stats] if class_stats else []

    return render_template('statistics.html',
                         expense_labels=json.dumps(expense_labels),
                         expense_values=json.dumps(expense_values),
                         income_labels=json.dumps(income_labels),
                         income_values=json.dumps(income_values),
                         class_labels=json.dumps(class_labels),
                         class_values=json.dumps(class_values),
                         colors=json.dumps(colors),
                         filter_period=filter_period)

# Wallet Blueprint
wallet_bp = Blueprint('wallet', __name__, url_prefix='/wallet')

@wallet_bp.route('/')
def index():
    """Halaman dompet pondok"""
    user_id = session.get('user_id', 1)

    wallet = query_db('SELECT balance FROM wallet WHERE user_id = ?', (user_id,), one=True)
    balance = wallet['balance'] if wallet else 0

    # Ambil transaksi terakhir
    transactions = query_db('''
        SELECT * FROM transactions
        WHERE user_id = ?
        ORDER BY date DESC, created_at DESC
        LIMIT 20
    ''', (user_id,))

    return render_template('wallet.html', balance=balance, transactions=transactions)

# Settings Blueprint
settings_bp = Blueprint('settings', __name__, url_prefix='/settings')

# Students Blueprint
students_bp = Blueprint('students', __name__, url_prefix='/students')

@settings_bp.route('/')
def index():
    """Halaman pengaturan"""
    user_id = session.get('user_id', 1)

    user = query_db('SELECT * FROM users WHERE id = ?', (user_id,), one=True)
    settings = query_db('SELECT `key`, `value` FROM settings')

    settings_dict = {row['key']: row['value'] for row in settings}

    return render_template('settings.html', user=user, settings=settings_dict)

@settings_bp.route('/update-profile', methods=['POST'])
@admin_required
def update_profile():
    """Update profil user"""
    user_id = session.get('user_id', 1)

    full_name = request.form.get('full_name')
    email = request.form.get('email')
    username = request.form.get('username')
    current_password = request.form.get('current_password')
    new_password = request.form.get('new_password')
    confirm_password = request.form.get('confirm_password')

    # Validasi username unik (kecuali untuk user sendiri)
    existing_user = get_user_by_username(username)
    if existing_user and existing_user['id'] != user_id:
        flash('Username sudah digunakan oleh user lain', 'danger')
        return redirect(url_for('settings.index'))

    # Validasi password jika diisi
    if new_password:
        if not current_password:
            
            return redirect(url_for('settings.index'))
        
        user = get_user(user_id)
        if not check_password_hash(user['password'], current_password):
            flash('Password lama salah', 'danger')
            return redirect(url_for('settings.index'))
        
        if new_password != confirm_password:
            flash('Konfirmasi password baru tidak cocok', 'danger')
            return redirect(url_for('settings.index'))
        
        if len(new_password) < 6:
            
            return redirect(url_for('settings.index'))

    # Update basic fields
    execute_db(
        'UPDATE users SET username = ?, full_name = ?, email = ? WHERE id = ?',
        (username, full_name, email, user_id)
    )
    
    # Update password jika diisi
    if new_password:
        set_user_password(user_id, new_password)
        flash('Password berhasil diubah', 'success')
        
    # Handle profile picture upload (optional)
    if 'profile_picture' in request.files:
        file = request.files['profile_picture']
        if file and file.filename:
            filename = secure_filename(file.filename)
            ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
            allowed = {'png', 'jpg', 'jpeg', 'gif'}
            if ext in allowed:
                try:
                    # Use current_app.root_path for correct path
                    upload_dir = os.path.join(current_app.root_path, 'static', 'uploads')
                    if not os.path.exists(upload_dir):
                        os.makedirs(upload_dir, exist_ok=True)
                    new_name = f"user_{user_id}_{int(datetime.now().timestamp())}.{ext}"
                    save_path = os.path.join(upload_dir, new_name)
                    file.save(save_path)
                    # Store relative path under static (e.g. uploads/filename)
                    picture_db_path = f"uploads/{new_name}"
                    update_user_profile_picture(user_id, picture_db_path)
                    session['profile_picture'] = picture_db_path
                    flash('Foto profil berhasil diperbarui', 'success')
                except Exception as e:
                    flash(f'Gagal mengupload foto profil: {str(e)}', 'danger')
            else:
                flash('Format file tidak didukung. Gunakan PNG, JPG, JPEG, atau GIF.', 'danger')
                
    session['full_name'] = full_name

    try:
        record_history(user_id, 'update', 'user', user_id, full_name)
    except Exception:
        pass

    flash('Profil berhasil diperbarui', 'success')
    return redirect(url_for('settings.index'))

@students_bp.route('/')
def index():
    """Halaman daftar santri"""
    students_data = get_all_students()
    
    # Konvert sqlite3.Row ke dictionary dan tambah statistik
    students = []
    for student in students_data:
        student_dict = dict(student)
        stats = get_student_payment_stats(student['id'])
        student_dict['total_payment'] = stats['total_payment']
        student_dict['month_payment'] = stats['month_payment']
        students.append(student_dict)
    
    return render_template('students.html', students=students)

@students_bp.route('/<int:student_id>')
def detail(student_id):
    """Detail santri dan pembayaran"""
    student = get_student(student_id)
    if not student:
        return redirect(url_for('students.index'))
    
    payments = get_student_payments(student_id)
    stats = get_student_payment_stats(student_id)
    
    return render_template('student_detail.html', student=student, payments=payments, stats=stats)

@students_bp.route('/add', methods=['GET', 'POST'])
def add():
    """Tambah santri baru"""
    if request.method == 'POST':
        name = request.form.get('name')
        nisn = request.form.get('nisn')
        kelas = request.form.get('kelas')
        jenis_kelamin = request.form.get('jenis_kelamin')
        phone = request.form.get('phone')
        parent_name = request.form.get('parent_name')
        parent_phone = request.form.get('parent_phone')
        alamat = request.form.get('alamat')
        
        new_id = add_student(name, nisn, kelas, jenis_kelamin, phone, parent_name, parent_phone, alamat)
        try:
            record_history(session.get('user_id'), 'create', 'student', new_id, name)
        except Exception:
            pass
        return redirect(url_for('students.index'))
    
    return render_template('add_student.html')

@students_bp.route('/<int:student_id>/edit', methods=['GET', 'POST'])
def edit(student_id):
    """Edit data santri"""
    student = get_student(student_id)
    if not student:
        return redirect(url_for('students.index'))
    
    if request.method == 'POST':
        name = request.form.get('name')
        nisn = request.form.get('nisn')
        kelas = request.form.get('kelas')
        jenis_kelamin = request.form.get('jenis_kelamin')
        phone = request.form.get('phone')
        parent_name = request.form.get('parent_name')
        parent_phone = request.form.get('parent_phone')
        alamat = request.form.get('alamat')
        status = request.form.get('status')
        # Update basic fields
        update_student(student_id, name, nisn, kelas, jenis_kelamin, phone, parent_name, parent_phone, alamat, status)

        # Handle photo upload (optional)
        if 'photo' in request.files:
            file = request.files.get('photo')
            if file and file.filename:
                filename = secure_filename(file.filename)
                ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
                allowed = {'png', 'jpg', 'jpeg', 'gif'}
                if ext in allowed:
                    upload_dir = os.path.join(os.path.dirname(__file__), 'static', 'uploads')
                    if not os.path.exists(upload_dir):
                        os.makedirs(upload_dir, exist_ok=True)
                    new_name = f"student_{student_id}_{int(datetime.now().timestamp())}.{ext}"
                    save_path = os.path.join(upload_dir, new_name)
                    file.save(save_path)
                    # Store relative path under static (e.g. uploads/filename)
                    photo_db_path = f"uploads/{new_name}"
                    try:
                        from db import update_student_photo
                        update_student_photo(student_id, photo_db_path)
                    except Exception:
                        # ignore failure to update photo in DB
                        pass

        # Prepare changes for history
        old_data = dict(student)
        new_data = {
            'name': name, 'nisn': nisn, 'kelas': kelas, 'jenis_kelamin': jenis_kelamin,
            'phone': phone, 'parent_name': parent_name, 'parent_phone': parent_phone,
            'alamat': alamat, 'status': status
        }
        changes = {k: v for k, v in new_data.items() if str(v) != str(old_data.get(k))}
        meta_info = json.dumps(changes) if changes else name

        try:
            record_history(session.get('user_id'), 'update', 'student', student_id, meta_info)
        except Exception:
            pass

        flash('Data santri berhasil diperbarui', 'success')
        return redirect(url_for('students.detail', student_id=student_id))
    # Convert to dict for template convenience
    try:
        student = dict(student)
    except Exception:
        pass

    return render_template('edit_student.html', student=student)

@students_bp.route('/<int:student_id>/delete', methods=['POST'])
def delete(student_id):
    """Hapus santri"""
    delete_student(student_id)
    try:
        record_history(session.get('user_id'), 'delete', 'student', student_id, None)
    except Exception:
        pass
    return redirect(url_for('students.index'))

@students_bp.route('/<int:student_id>/add-payment', methods=['POST'])
def add_payment(student_id):
    """Tambah pembayaran untuk santri"""
    amount = int(request.form.get('amount', 0))
    description = request.form.get('description', 'Pembayaran Santri')
    date_str = request.form.get('date', datetime.now().strftime('%Y-%m-%d'))
    
    user_id = session.get('user_id', 1)
    
    # Insert transaksi
    execute_db('''
        INSERT INTO transactions (user_id, student_id, type, category, amount, description, date)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (user_id, student_id, 'income', 'Pembayaran Santri', amount, description, date_str))

    # Update balance
    wallet = query_db('SELECT balance FROM wallet WHERE user_id = ?', (user_id,), one=True)
    new_balance = wallet['balance'] + amount
    execute_db('UPDATE wallet SET balance = ? WHERE user_id = ?', (new_balance, user_id))
    
    return redirect(url_for('students.detail', student_id=student_id))


@students_bp.route('/export-excel')
def export_excel():
    """Export data santri ke Excel"""
    students_data = get_all_students()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Santri"
    
    # Define styles
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add headers
    headers = ['No.', 'Nama Santri', 'NISN', 'Kelas', 'Jenis Kelamin', 'No. HP', 'Nama Orang Tua', 'No. HP Orang Tua', 'Alamat', 'Status']
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Add data
    for idx, student in enumerate(students_data, 1):
        ws.append([
            idx,
            student['name'],
            student['nisn'],
            student['kelas'],
            student['jenis_kelamin'],
            student['phone'],
            student['parent_name'],
            student['parent_phone'],
            student['alamat'],
            student['status']
        ])
    
    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 25
    ws.column_dimensions['J'].width = 12
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'data_santri_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )


@students_bp.route('/import-excel', methods=['GET', 'POST'])
def import_excel():
    """Import data santri dari Excel dengan validasi duplikasi"""
    if request.method == 'POST':
        if 'file' not in request.files:
            return redirect(url_for('students.index'))
        
        file = request.files['file']
        if file.filename == '':
            return redirect(url_for('students.index'))
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return render_template('import_students.html', error="File harus berformat Excel (.xlsx atau .xls)")
        
        try:
            # Load workbook
            wb = load_workbook(file)
            ws = wb.active
            
            # Collect existing NISN to check duplicates
            existing_nisns = set(row['nisn'] for row in get_all_students() if row['nisn'])
            existing_names = {}
            for row in get_all_students():
                key = (row['name'].lower().strip(), row['kelas'])
                existing_names[key] = row['id']
            
            imported = 0
            duplicates = []
            errors = []
            
            # Process rows (skip header)
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if not row[1]:  # Skip empty rows
                    continue
                
                try:
                    name = str(row[1]).strip() if row[1] else ''
                    nisn = str(row[2]).strip() if row[2] else ''
                    kelas = str(row[3]).strip() if row[3] else ''
                    jenis_kelamin = str(row[4]).strip() if row[4] else 'Laki-laki'
                    phone = str(row[5]).strip() if row[5] else ''
                    parent_name = str(row[6]).strip() if row[6] else ''
                    parent_phone = str(row[7]).strip() if row[7] else ''
                    alamat = str(row[8]).strip() if row[8] else ''
                    status = str(row[9]).strip() if row[9] else 'aktif'
                    
                    # Validasi
                    if not name:
                        errors.append(f"Baris {row_idx}: Nama santri tidak boleh kosong")
                        continue
                    
                    # Check for exact duplicate (same name in same class)
                    name_key = (name.lower(), kelas)
                    if name_key in existing_names:
                        duplicates.append({
                            'name': name,
                            'kelas': kelas,
                            'nisn': nisn,
                            'reason': 'Sudah ada santri dengan nama yang sama di kelas ini'
                        })
                        continue
                    
                    # Check if NISN already exists (if NISN provided)
                    if nisn and nisn in existing_nisns:
                        duplicates.append({
                            'name': name,
                            'kelas': kelas,
                            'nisn': nisn,
                            'reason': 'NISN sudah terdaftar'
                        })
                        continue
                    
                    # Add student
                    add_student(name, nisn, kelas, jenis_kelamin, phone, parent_name, parent_phone, alamat, status)
                    existing_nisns.add(nisn)
                    existing_names[name_key] = True
                    imported += 1
                    
                except Exception as e:
                    errors.append(f"Baris {row_idx}: {str(e)}")
            
            return render_template('import_students.html', 
                                 success_message=f"Berhasil import {imported} santri",
                                 duplicates=duplicates,
                                 errors=errors)
        
        except Exception as e:
            return render_template('import_students.html', error=f"Error membaca file: {str(e)}")
    
    return render_template('import_students.html')


@students_bp.route('/download-report')
def download_report():
    """Download laporan data santri dengan statistik pembayaran"""
    students_data = get_all_students()
    
    # Create workbook
    wb = Workbook()
    
    # Sheet 1: Data Santri dengan Statistik
    ws1 = wb.active
    ws1.title = "Laporan Santri"
    
    # Header styling
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws1['A1'] = "LAPORAN DATA SANTRI"
    ws1['A1'].font = Font(bold=True, size=14, color="6366F1")
    ws1.merge_cells('A1:K1')
    ws1['A2'] = f"Tanggal: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}"
    ws1.merge_cells('A2:K2')
    
    # Headers
    headers = ['No.', 'Nama Santri', 'NISN', 'Kelas', 'Jenis Kelamin', 'No. HP', 'Nama Orang Tua', 
               'No. HP Orang Tua', 'Status', 'Total Pembayaran', 'Pembayaran Bulan Ini']
    ws1.append([])  # Blank row
    ws1.append(headers)
    
    # Style headers
    for cell in ws1[4]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Add data with payment stats
    for idx, student in enumerate(students_data, 1):
        stats = get_student_payment_stats(student['id'])
        ws1.append([
            idx,
            student['name'],
            student['nisn'],
            student['kelas'],
            student['jenis_kelamin'],
            student['phone'],
            student['parent_name'],
            student['parent_phone'],
            student['status'],
            stats['total_payment'],
            stats['month_payment']
        ])
    
    # Set column widths
    widths = [5, 20, 15, 12, 15, 15, 20, 15, 12, 15, 15]
    for idx, width in enumerate(widths, 1):
        ws1.column_dimensions[chr(64 + idx)].width = width
    
    # Sheet 2: Ringkasan Statistik
    ws2 = wb.create_sheet("Statistik")
    
    total_students = len(students_data)
    aktif_count = sum(1 for s in students_data if s['status'] == 'aktif')
    non_aktif_count = total_students - aktif_count
    
    # Hitung total pembayaran
    total_payment = sum(get_student_payment_stats(s['id'])['total_payment'] for s in students_data)
    belum_bayar_count = sum(1 for s in students_data if get_student_payment_stats(s['id'])['month_payment'] == 0)
    
    ws2['A1'] = "RINGKASAN STATISTIK"
    ws2['A1'].font = Font(bold=True, size=14, color="6366F1")
    ws2.merge_cells('A1:B1')
    
    # Statistics
    stats_data = [
        ('Total Santri', total_students),
        ('Santri Aktif', aktif_count),
        ('Santri Non-Aktif', non_aktif_count),
        ('Total Pembayaran', f"Rp {total_payment:,.0f}"),
        ('Belum Bayar Bulan Ini', belum_bayar_count),
    ]
    
    ws2.append([])
    for label, value in stats_data:
        ws2.append([label, value])
    
    # Style statistics
    for row in ws2.iter_rows(min_row=3, max_row=7):
        row[0].font = Font(bold=True)
        row[0].fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
        row[1].border = border
    
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 20
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'laporan_santri_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )


# History / Activity Log Blueprint
history_bp = Blueprint('history', __name__, url_prefix='/history')


@history_bp.route('/')
def index():
    """Tampilkan riwayat aktivitas"""
    # Optionally limit or paginate
    entries = get_history(500)
    entries = [dict(e) for e in entries] if entries else []
    return render_template('history.html', entries=entries)


@history_bp.route('/delete/<int:entry_id>', methods=['POST'])
def delete(entry_id):
    """Hapus satu entri history"""
    execute_db('DELETE FROM history WHERE id = ?', (entry_id,))
    return redirect(url_for('history.index'))


# --- User management (admin only) ---
users_bp = Blueprint('users', __name__, url_prefix='/users')




@users_bp.route('/')
@admin_required
def index():
    users = get_all_users()
    users = [dict(u) for u in users] if users else []
    return render_template('users.html', users=users)


@users_bp.route('/create', methods=['GET', 'POST'])
@admin_required
def create():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        email = request.form.get('email', '')
        full_name = request.form.get('full_name', '')
        role = request.form.get('role', 'user')
        create_user(username, password, email, full_name, role)
        try:
            record_history(session.get('user_id'), 'create', 'user', None, username)
        except Exception:
            pass
        return redirect(url_for('users.index'))
    return render_template('user_form.html', user=None)


@users_bp.route('/edit/<int:user_id>', methods=['GET', 'POST'])
@admin_required
def edit(user_id):
    user = get_user(user_id)
    if not user:
        return redirect(url_for('users.index'))
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email', '')
        full_name = request.form.get('full_name', '')
        role = request.form.get('role', 'user')
        update_user(user_id, username, email, full_name, role)
        # Optional password change
        new_pw = request.form.get('password')
        if new_pw:
            set_user_password(user_id, new_pw)
        try:
            record_history(session.get('user_id'), 'update', 'user', user_id, username)
        except Exception:
            pass
        return redirect(url_for('users.index'))
    try:
        user = dict(user)
    except Exception:
        pass
    return render_template('user_form.html', user=user)


@users_bp.route('/delete/<int:user_id>', methods=['POST'])
@admin_required
def delete(user_id):
    delete_user(user_id)
    try:
        record_history(session.get('user_id'), 'delete', 'user', user_id, None)
    except Exception:
        pass
    return redirect(url_for('users.index'))


# --- Pembayaran Santri (placeholder) ---
payments_bp = Blueprint('payments', __name__, url_prefix='/payments')

@payments_bp.route('/')
def index_payments():
    # Visible to admin and staff
    if session.get('role') not in ('admin', 'staff'):
        return redirect(url_for('dashboard.index'))
    # List bills
    bills_data = get_all_bills()
    bills = []
    from db import get_bill_total_paid
    for b in bills_data:
        b_dict = dict(b)
        b_dict['paid_amount'] = get_bill_total_paid(b['id'])
        bills.append(b_dict)
    return render_template('payments_list.html', bills=bills)



@payments_bp.route('/create', methods=['GET', 'POST'])
def create_bill_view():
    if session.get('role') not in ('admin', 'staff'):
        return redirect(url_for('dashboard.index'))
    if request.method == 'POST':
        title = request.form.get('title')
        amount = int(request.form.get('amount', 0))
        due_date = request.form.get('due_date') or None

        # Support multiple student selection via checkboxes (student_ids)
        student_ids = request.form.getlist('student_ids')  # may be []

        created = []
        if student_ids:
            for sid in student_ids:
                try:
                    sid_int = int(sid)
                except Exception:
                    continue
                b_id = create_bill(sid_int, title, amount, due_date, session.get('user_id'))
                created.append(b_id)
        else:
            # Fallback to single student_id (old behavior)
            student_id = request.form.get('student_id')
            if student_id:
                try:
                    sid_int = int(student_id)
                    b_id = create_bill(sid_int, title, amount, due_date, session.get('user_id'))
                    created.append(b_id)
                except Exception:
                    pass

        # Record history for created bills
        try:
            for b in created:
                record_history(session.get('user_id'), 'create', 'bill', b, f"{title}:{amount}")
        except Exception:
            pass

        if created:
            flash(f"Berhasil membuat {len(created)} tagihan", 'success')
        else:
            flash('Tidak ada tagihan dibuat. Pastikan Anda memilih santri.', 'warning')

        return redirect(url_for('payments.index_payments'))

    # GET -> form
    students = get_all_students()
    students = [dict(s) for s in students] if students else []
    
    # Add unpaid amount for each student
    for student in students:
        student['unpaid_amount'] = get_student_unpaid_amount(student['id'])
    
    return render_template('payment_form.html', students=students, bill=None)


@payments_bp.route('/edit/<int:bill_id>', methods=['GET', 'POST'])
def edit_bill_view(bill_id):
    if session.get('role') not in ('admin', 'staff'):
        return redirect(url_for('dashboard.index'))
    bill = get_bill(bill_id)
    if not bill:
        return redirect(url_for('payments.index_payments'))
    if request.method == 'POST':
        student_id = int(request.form.get('student_id'))
        title = request.form.get('title')
        amount = int(request.form.get('amount', 0))
        due_date = request.form.get('due_date') or None
        status = request.form.get('status', 'unpaid')
        update_bill(bill_id, student_id, title, amount, due_date, status)
        try:
            record_history(session.get('user_id'), 'update', 'bill', bill_id, f"{title}:{amount}")
        except Exception:
            pass
        return redirect(url_for('payments.index_payments'))

    students_data = get_all_students()
    students = []
    for s in students_data:
        s_dict = dict(s)
        s_dict['unpaid_amount'] = get_student_unpaid_amount(s['id'])
        students.append(s_dict)

    try:
        bill = dict(bill)
    except Exception:
        pass
    return render_template('payment_form.html', students=students, bill=bill)


@payments_bp.route('/delete/<int:bill_id>', methods=['POST'])
def delete_bill_view(bill_id):
    if session.get('role') not in ('admin', 'staff'):
        return redirect(url_for('dashboard.index'))
    delete_bill(bill_id)
    try:
        record_history(session.get('user_id'), 'delete', 'bill', bill_id, None)
    except Exception:
        pass
    return redirect(url_for('payments.index_payments'))


@payments_bp.route('/pay/<int:bill_id>', methods=['POST'])
def pay_bill_view(bill_id):
    if session.get('role') not in ('admin', 'staff'):
        return redirect(url_for('dashboard.index'))
    bill = get_bill(bill_id)
    if not bill:
        return redirect(url_for('payments.index_payments'))

    try:
        amount_to_pay = int(request.form.get('amount', 0))
        if amount_to_pay <= 0:
            flash('Jumlah pembayaran harus lebih dari 0', 'warning')
            return redirect(url_for('payments.index_payments'))
            
        user_id = session.get('user_id', 1)
        student_id = bill['student_id']
        title = bill['title']
        today = datetime.now().strftime('%Y-%m-%d')

        # Insert transaction (income) linked to bill
        execute_db(
            '''INSERT INTO transactions (user_id, student_id, type, category, amount, description, date, bill_id)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
            (user_id, student_id, 'income', 'Pembayaran Santri', amount_to_pay, title, today, bill_id)
        )

        # Update wallet balance
        wallet = query_db('SELECT balance FROM wallet WHERE user_id = ?', (user_id,), one=True)
        current_balance = wallet['balance'] if wallet else 0
        new_balance = current_balance + amount_to_pay
        execute_db('UPDATE wallet SET balance = ? WHERE user_id = ?', (new_balance, user_id))

        # Check if bill is now fully paid
        from db import get_bill_total_paid
        total_paid = get_bill_total_paid(bill_id)
        if total_paid >= bill['amount']:
            mark_bill_paid(bill_id)
            flash(f"Pembayaran Rp {amount_to_pay:,.0f} berhasil. Tagihan '{title}' sekarang Lunas.", 'success')
        else:
            flash(f"Pembayaran Rp {amount_to_pay:,.0f} berhasil. Sisa tagihan: Rp {(bill['amount'] - total_paid):,.0f}", 'success')

        record_history(user_id, 'pay', 'bill', bill_id, f"{title}:{amount_to_pay}")
    except Exception as e:
        flash(f'Terjadi masalah saat memproses pembayaran: {str(e)}', 'danger')

    return redirect(url_for('payments.index_payments'))


@payments_bp.route('/receipt/<int:bill_id>')
def bill_receipt(bill_id):
    """Lookup transaction for a bill and redirect to receipt"""
    bill = get_bill(bill_id)
    if not bill or bill['status'] != 'paid':
        flash('Tagihan belum dibayar atau tidak ditemukan', 'warning')
        return redirect(url_for('payments.index_payments'))
    
    # Try to find the latest transaction that matches this bill
    # We use amount and description (title) and student_id
    query = '''
        SELECT id FROM transactions 
        WHERE student_id = ? AND amount = ? AND (description = ? OR description LIKE ?)
        AND type = 'income'
        ORDER BY created_at DESC LIMIT 1
    '''
    trans = query_db(query, (bill['student_id'], bill['amount'], bill['title'], f"%{bill['title']}%"), one=True)
    
    if trans:
        return redirect(url_for('transaction.receipt', id=trans['id']))
    else:
        flash('Kuitansi transaksi tidak ditemukan untuk tagihan ini', 'danger')
        return redirect(url_for('payments.index_payments'))


